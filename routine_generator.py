import pandas as pd
from datetime import datetime, timedelta
import random
import openpyxl

class RoutineGenerator:
    def __init__(self, working_days=None, periods_per_day=6, time_slots=None):
        self.periods_per_day = periods_per_day
        self.days = working_days if working_days else ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday']
        
        # Generate time slots if not provided
        if time_slots:
            self.time_slots = time_slots
        else:
            # Default time slots from 8:30 AM with 1-hour intervals
            start_time = datetime.strptime("8:30", "%H:%M")
            self.time_slots = []
            for i in range(self.periods_per_day):
                end_time = start_time + timedelta(hours=1)
                self.time_slots.append(f"{start_time.strftime('%I:%M')}-{end_time.strftime('%I:%M')}")
                start_time = end_time
        
    def generate_routine(self, classes, teachers, subjects):
        """
        Generate routines for multiple classes ensuring no teacher conflicts
        
        Args:
            classes (list): List of class names
            teachers (dict): Dictionary mapping subjects to teachers
            subjects (dict): Dictionary mapping classes to their subjects
        """
        all_routines = {}
        teacher_schedule = {day: {slot: [] for slot in self.time_slots} 
                          for day in self.days}
        
        for class_name in classes:
            routine = {day: {slot: '' for slot in self.time_slots} 
                      for day in self.days}
            class_subjects = subjects[class_name]
            
            for day in self.days:
                available_slots = list(self.time_slots)
                random.shuffle(class_subjects)
                
                for subject in class_subjects:
                    if not available_slots:
                        break
                        
                    teacher = random.choice(teachers[subject])
                    
                    # Find a slot where the teacher is available
                    valid_slots = [
                        slot for slot in available_slots 
                        if teacher not in teacher_schedule[day][slot]
                    ]
                    
                    if valid_slots:
                        slot = random.choice(valid_slots)
                        routine[day][slot] = f"{subject}\n({teacher})"
                        teacher_schedule[day][slot].append(teacher)
                        available_slots.remove(slot)
            
            all_routines[class_name] = routine
        
        return all_routines
    
    def save_to_excel(self, routines, output_file):
        """Save generated routines to an Excel file with multiple sheets"""
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            for class_name, routine in routines.items():
                # Create DataFrame
                df = pd.DataFrame(routine)
                df.index = self.time_slots
                
                # Write DataFrame to Excel
                df.to_excel(writer, sheet_name=f'Class {class_name}')
                
                # Get the worksheet
                worksheet = writer.sheets[f'Class {class_name}']
                
                # Get workbook
                workbook = writer.book
                
                # Define styles
                header_fill = openpyxl.styles.PatternFill(
                    start_color='1F4E78',
                    end_color='1F4E78',
                    fill_type='solid'
                )
                header_font = openpyxl.styles.Font(
                    color='FFFFFF',
                    bold=True,
                    size=12
                )
                time_fill = openpyxl.styles.PatternFill(
                    start_color='D9E1F2',
                    end_color='D9E1F2',
                    fill_type='solid'
                )
                time_font = openpyxl.styles.Font(
                    bold=True,
                    size=11
                )
                cell_border = openpyxl.styles.Border(
                    left=openpyxl.styles.Side(style='thin'),
                    right=openpyxl.styles.Side(style='thin'),
                    top=openpyxl.styles.Side(style='thin'),
                    bottom=openpyxl.styles.Side(style='thin')
                )
                
                # Set column widths
                worksheet.column_dimensions['A'].width = 20  # Time slot column
                
                # Set wider width for all other columns
                for col in worksheet.columns:
                    column = col[0].column_letter
                    if column != 'A':  # Skip the first column as it's already set
                        worksheet.column_dimensions[column].width = 30  # Fixed wider width for day columns
                
                # Set row heights for better vertical spacing
                for row in worksheet.rows:
                    max_lines = 1
                    for cell in row:
                        if cell.value:
                            max_lines = max(max_lines, str(cell.value).count('\n') + 1)
                    worksheet.row_dimensions[row[0].row].height = max_lines * 25  # Increased height multiplier
                
                # Apply styles to headers (days)
                for col in range(2, len(self.days) + 2):
                    cell = worksheet.cell(row=1, column=col)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.border = cell_border
                    cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
                
                # Apply styles to time slots
                for row in range(2, len(self.time_slots) + 2):
                    cell = worksheet.cell(row=row, column=1)
                    cell.fill = time_fill
                    cell.font = time_font
                    cell.border = cell_border
                    cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
                
                # Apply styles to all cells
                for row in range(2, len(self.time_slots) + 2):
                    for col in range(2, len(self.days) + 2):
                        cell = worksheet.cell(row=row, column=col)
                        cell.border = cell_border
                        cell.alignment = openpyxl.styles.Alignment(
                            horizontal='center',
                            vertical='center',
                            wrap_text=True
                        )
                        
                        # Add alternating row colors
                        if row % 2 == 0:
                            cell.fill = openpyxl.styles.PatternFill(
                                start_color='F5F5F5',
                                end_color='F5F5F5',
                                fill_type='solid'
                            )
                
                # Add title
                title_row = 1
                worksheet.insert_rows(1)
                title_cell = worksheet.cell(row=title_row, column=1)
                title_cell.value = f"Class {class_name} - Routine"
                title_cell.font = openpyxl.styles.Font(size=14, bold=True, color='1F4E78')
                worksheet.merge_cells(start_row=title_row, start_column=1,
                                   end_row=title_row, end_column=len(self.days) + 1)
                title_cell.alignment = openpyxl.styles.Alignment(horizontal='center')
                
                
                # Freeze panes
                worksheet.freeze_panes = 'B3'

def get_teacher_info():
    """Get teacher information from user input"""
    teachers = {}
    while True:
        teacher_name = input("Enter teacher name (or 'done' to finish): ").strip()
        if teacher_name.lower() == 'done':
            break
        
        subjects = []
        print(f"Enter subjects for {teacher_name} (type 'done' when finished):")
        while True:
            subject = input("Enter subject: ").strip()
            if subject.lower() == 'done':
                break
            subjects.append(subject)
        
        for subject in subjects:
            if subject not in teachers:
                teachers[subject] = []
            teachers[subject].append(teacher_name)
    
    return teachers

def get_class_info():
    """Get class information from user input"""
    classes = []
    subjects = {}
    
    while True:
        class_name = input("Enter class name (or 'done' to finish): ").strip()
        if class_name.lower() == 'done':
            break
        
        classes.append(class_name)
        class_subjects = []
        print(f"Enter subjects for class {class_name} (type 'done' when finished):")
        while True:
            subject = input("Enter subject: ").strip()
            if subject.lower() == 'done':
                break
            class_subjects.append(subject)
        
        subjects[class_name] = class_subjects
    
    return classes, subjects

def main():
    print("Welcome to Class Routine Generator!")
    print("\nFirst, let's get information about the teachers and their subjects.")
    teachers = get_teacher_info()
    
    print("\nNow, let's get information about the classes and their subjects.")
    classes, subjects = get_class_info()
    
    # Validate that all subjects have teachers
    all_subjects = set()
    for class_subjects in subjects.values():
        all_subjects.update(class_subjects)
    
    missing_subjects = [subject for subject in all_subjects if subject not in teachers]
    if missing_subjects:
        print("\nError: The following subjects don't have assigned teachers:")
        for subject in missing_subjects:
            print(f"- {subject}")
        return
    
    # Get custom days and periods from user
    working_days = input("\nEnter working days (comma-separated, e.g., Monday, Tuesday, Wednesday): ").strip().split(',')
    working_days = [day.strip() for day in working_days]
    
    periods_per_day = int(input("Enter number of periods per day: "))
    
    generator = RoutineGenerator(working_days, periods_per_day)
    routines = generator.generate_routine(classes, teachers, subjects)
    generator.save_to_excel(routines, 'class_routines.xlsx')
    print("\nClass routines have been generated and saved to 'class_routines.xlsx'")

if __name__ == "__main__":
    main()
